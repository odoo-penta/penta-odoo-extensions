# wizards/lc_taxes_io_wizard.py
# -*- coding: utf-8 -*-
import base64
import io
from odoo import api, fields, models, _
from odoo.exceptions import UserError


EXPORT_HEADERS = [
    'id_line',             # 0. id de la línea
    'Impuesto',               # 1. Cost.line.id (Nombre)
    'Producto',                 # 2. Product.id (Nombre)
    'Cantidad',                # 3. Quantity
    'Impuesto adicional',  # 4. Tax calculado
]

class LCTaxesIoWizard(models.TransientModel):
    _name = 'lc.taxes.io.wizard'
    _description = 'Exportar impuestos de Costo en Destino'

    export_file = fields.Binary('Archivo a descargar', readonly=True)
    export_filename = fields.Char('Nombre de archivo', readonly=True)

    import_file = fields.Binary('Archivo a importar (XLSX)')
    import_filename = fields.Char('Nombre del archivo')
    
    
    def _get_landed_cost(self):
        active_id = self.env.context.get('active_id')
        if not active_id:
            raise UserError(_("Este asistente debe abrirse desde un Costo en Destino."))
        lc = self.env['stock.landed.cost'].browse(active_id)
        if not lc.exists():
            raise UserError(_("No se encontró el Costo en Destino activo."))
        return lc

    def action_confirm_import(self):
        """Importa desde XLSX (hoja 'taxes') con columnas:
        tax_line_id, cost_line, product, quantity, additional_landed_cost
        Solo ACTUALIZA: quantity y tax_value.
        """
        if not self.import_file:
            raise UserError(_("Sube un archivo XLSX para importar."))

        lc = self._get_landed_cost()

        try:
            import openpyxl
        except ImportError:
            raise UserError(_("Falta la librería 'openpyxl' en el servidor."))

        content = base64.b64decode(self.import_file or b'')
        wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)

        if 'taxes' not in wb.sheetnames:
            raise UserError(_("La hoja 'taxes' no existe en el archivo."))

        ws = wb['taxes']

        # Mapeo de cabeceras
        headers = {}
        for c, cell in enumerate(ws[1], start=1):
            key = (cell.value or '').strip()
            if key:
                headers[key] = c

        # Requeridas para actualizar
        required = {'tax_line_id', 'quantity', 'additional_landed_cost'}
        missing = [h for h in required if h not in headers]
        if missing:
            raise UserError(_("Faltan columnas obligatorias: %s") % ', '.join(missing))

        TaxLine = self.env['stock.landed.cost.taxes']
        updated, errors = 0, []

        for r in range(2, ws.max_row + 1):
            try:
                tlid_cell = ws.cell(row=r, column=headers['tax_line_id']).value
                if not tlid_cell:
                    errors.append(f"Fila {r}: Falta tax_line_id.")
                    continue
                try:
                    tlid = int(tlid_cell)
                except Exception:
                    errors.append(f"Fila {r}: tax_line_id inválido: {tlid_cell}")
                    continue

                line = TaxLine.browse(tlid)
                if not line or not line.exists():
                    errors.append(f"Fila {r}: Línea de impuesto (ID {tlid}) no existe.")
                    continue

                # Debe pertenecer al LC activo
                if line.landed_cost_id.id != lc.id:
                    errors.append(f"Fila {r}: Línea {tlid} no pertenece al Costo en Destino {lc.id}.")
                    continue

                # Lee valores a actualizar
                qty_cell = ws.cell(row=r, column=headers['quantity']).value
                add_cost_cell = ws.cell(row=r, column=headers['additional_landed_cost']).value

                qty = float(qty_cell or 0.0)
                tax_val = float(add_cost_cell or 0.0)

                line.write({
                    'quantity': qty,
                    'tax_value': tax_val,
                })
                updated += 1

            except Exception as e:
                errors.append(f"Fila {r}: {e}")

        if errors:
            detalle = "\n- " + "\n- ".join(errors[:50])  # limita si quieres
            raise UserError(_("Errores durante la importación:\n%s") % detalle)
        
        return {'type': 'ir.actions.act_window_close'}

        return {
            'type': 'ir.actions.act_window',
            'res_model': 'lc.taxes.io.wizard',
            'res_id': self.id,
            'view_mode': 'form',
            'target': 'new',
            'context': self.env.context,
        }


    def action_generate_xlsx(self):
        lc = self._get_landed_cost()
        TaxLine = self.env['stock.landed.cost.taxes']
        lines = TaxLine.search([('landed_cost_id', '=', lc.id)])

        try:
            import xlsxwriter
        except ImportError:
            raise UserError(_("Falta la librería 'xlsxwriter' en el servidor."))

        bio = io.BytesIO()
        wb = xlsxwriter.Workbook(bio, {'in_memory': True})
        ws = wb.add_worksheet('taxes')

        fmt_bold = wb.add_format({'bold': True})
        fmt_num = wb.add_format({'num_format': '#,##0.00'})

        # Cabeceras
        for c, head in enumerate(EXPORT_HEADERS):
            ws.write(0, c, head, fmt_bold)

        # Datos
        row = 1
        for tl in lines:
            vals = [
                tl.id,                                # 0. id de la línea
                tl.cost_line or '',                   # 1. cost_line_id (nombre texto)
                tl.product_id.display_name or '',     # 2. product_id (nombre)
                tl.quantity or 0.0,                   # 3. quantity
                tl.tax_value or 0.0,                  # 4. valor del impuesto calculado
            ]
            for c, v in enumerate(vals):
                if isinstance(v, (int, float)):
                    ws.write_number(row, c, v, fmt_num)
                else:
                    ws.write(row, c, v)
            row += 1

        wb.close()
        bio.seek(0)

        self.write({
            'export_file': base64.b64encode(bio.read()),
            'export_filename': f'landed_cost_taxes_{lc.id}.xlsx',
        })

        # 🔁 recargar wizard para mostrar el archivo
        return {
            'type': 'ir.actions.act_window',
            'res_model': 'lc.taxes.io.wizard',
            'res_id': self.id,
            'view_mode': 'form',
            'target': 'new',
            'context': self.env.context,
        }
