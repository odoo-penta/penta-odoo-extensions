# -*- coding: utf-8 -*-
from odoo import models, fields
from odoo.exceptions import UserError
import base64
import io
from openpyxl import load_workbook


class ImportPickingOperationsWizard(models.TransientModel):
    _name = 'import.picking.operations.wizard'
    _description = 'Wizard to import picking operations from an Excel file'

    file = fields.Binary(string="File Excel", required=True)
    filename = fields.Char(string="File name")

    def action_process_file(self):
        self.ensure_one()
        if not self.file:
            raise UserError("You must upload an Excel file.")
        try:
            # Decodificar el archivo
            file_content = base64.b64decode(self.file)
            file_stream = io.BytesIO(file_content)
            # Abrir el archivo Excel
            wb = load_workbook(filename=file_stream, read_only=True)
            sheet = wb.active  # primera hoja
            rows = []
            for i, row in enumerate(sheet.iter_rows(values_only=True), start=1):
                if i == 1:  # saltar cabecera
                    continue
                rows.append(row)
            picking_id = self.env.context.get('active_id', False)
            picking = self.env['stock.picking'].browse(picking_id)
                        
            for line in picking.move_ids_without_package:
                for move_line in line.move_line_ids:
                    for row in rows:
                        ref, product_name, serial, motor, ramv = row[0], row[1], row[2], row[3], row[4]

                        if move_line.product_id.default_code == ref and move_line.product_id.name == product_name:

                            # Buscar si ya existe lote
                            lot = self.env['stock.lot'].search([
                                ('name', '=', serial),
                                ('product_id', '=', move_line.product_id.id)
                            ], limit=1)

                            # Si no existe → crearlo
                            if not lot:
                                lot = self.env['stock.lot'].create({
                                    'name': serial,
                                    'product_id': move_line.product_id.id,
                                    'company_id': move_line.company_id.id,
                                    'motor_number': motor,
                                    'ramv': ramv,
                                })
                            else:
                                # Si existe → actualizar campos
                                lot.write({
                                    'motor_number': motor,
                                    'ramv': ramv,
                                })

                            # Asignar lote al move_line
                            move_line.lot_id = lot.id

                            # También llenar los campos del move_line
                            move_line.lot_name = serial
                            move_line.motor_number = motor
                            move_line.ramv = ramv

                            rows.remove(row)
                            break
        except Exception as e:
            # Si ocurre cualquier error lo mostramos en un popup
            raise UserError(f"Something went wrong while processing the file:\n{str(e)}")
        
        return {"type": "ir.actions.act_window_close"}